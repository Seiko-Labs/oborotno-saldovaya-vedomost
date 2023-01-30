import calendar
import datetime
import os
from typing import List, Tuple, Any, Dict, Set

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from data_structures import Dimension, BranchInfo


class Excel:
    def __init__(self, filename: str) -> None:
        self._workbook: Workbook = openpyxl.load_workbook(filename=filename, data_only=True)

        self._general_sheet: Worksheet = self._workbook.worksheets[0]
        self._branch_sheet: Worksheet = self._workbook.worksheets[1]

        self._general_dim: Dimension = Dimension(x=self._general_sheet.max_column, y=self._general_sheet.max_row)
        self._branch_dim: Dimension = Dimension(x=self._branch_sheet.max_column, y=self._branch_sheet.max_row)

        self._general_rows: List[Tuple[Any]] = list(self._general_sheet.iter_rows(3, self._general_dim.y, 2, self._general_dim.x, values_only=True))
        self._branch_cols: List[Tuple[Any]] = list(self._branch_sheet.iter_cols(2, self._branch_dim.x, 3, self._branch_dim.y, values_only=True))

        self.general_data: List[BranchInfo] = self._get_general_data()
        self.branch_data: List[BranchInfo] = self._get_branch_data()

    def _get_general_data(self) -> List[BranchInfo]:
        return [BranchInfo(branch='00', account=cells[1], account_name=cells[0]) for cells in self._general_rows if any(cells)]

    def _get_branch_data(self) -> List[BranchInfo]:
        infos: List[BranchInfo] = []
        for col in self._branch_cols:
            for cell in col[1::]:
                if cell == '-':
                    continue
                infos.append(BranchInfo(branch=f'{int(cell):02}', account=col[0]))
        return infos

    @staticmethod
    def _get_last_full(iterable) -> int:
        sz: int = len(list(iterable))
        try:
            return next(j for j, cells in enumerate(iterable) if not any(cells))
        except StopIteration:
            return sz


class DataGetter:
    def __init__(self, _date: datetime.datetime = None) -> None:
        function_data: Dict = {1: {'last_daily': ['Z_160_GL_020', 'Z_160_GL_003'], 'monthly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_003', 'S_CLI_004']}, 2: {'last_daily': ['Z_160_GL_020', 'Z_160_GL_003'], 'monthly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_003', 'S_CLI_004']}, 3: {'last_daily': ['Z_160_GL_020', 'Z_160_GL_003'], 'monthly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_003', 'S_CLI_004'], 'quarterly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_013', 'S_CLI_014']}, 4: {'last_daily': ['Z_160_GL_020', 'Z_160_GL_003'], 'monthly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_003', 'S_CLI_004']}, 5: {'last_daily': ['Z_160_GL_020', 'Z_160_GL_003'], 'monthly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_003', 'S_CLI_004']}, 6: {'last_daily': ['Z_160_GL_020', 'Z_160_GL_003'], 'monthly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_003', 'S_CLI_004'], 'quarterly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_013', 'S_CLI_014'], 'six_monthly': ['Z_160_GL_020', 'Z_160_GL_003']}, 7: {'last_daily': ['Z_160_GL_020', 'Z_160_GL_003'], 'monthly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_003', 'S_CLI_004']}, 8: {'last_daily': ['Z_160_GL_020', 'Z_160_GL_003'], 'monthly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_003', 'S_CLI_004']}, 9: {'last_daily': ['Z_160_GL_020', 'Z_160_GL_003'], 'monthly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_003', 'S_CLI_004'], 'quarterly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_013', 'S_CLI_014'], 'nine_monthly': ['Z_160_GL_020', 'Z_160_GL_003']}, 10: {'last_daily': ['Z_160_GL_020', 'Z_160_GL_003'], 'monthly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_003', 'S_CLI_004']}, 11: {'last_daily': ['Z_160_GL_020', 'Z_160_GL_003'], 'monthly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_003', 'S_CLI_004']}, 12: {'last_daily': ['Z_160_GL_020', 'Z_160_GL_003'], 'monthly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_003', 'S_CLI_004'], 'quarterly': ['Z_160_GL_020', 'Z_160_GL_003', 'S_CLI_013', 'S_CLI_014'], 'six_monthly': ['Z_160_GL_020', 'Z_160_GL_003'], 'yearly': ['Z_160_GL_020', 'Z_160_GL_003']}}

        today: datetime.datetime = _date if _date else datetime.datetime.now()
        month: int = today.month - 1
        year: int = today.year
        if today.month == 1:
            month = 12
            year = today.year - 1

        self.info: List[BranchInfo] = []

        excel: Excel = Excel(filename=rf'C:\Users\{os.getlogin()}\Desktop\ОСВ_филиалы_счета.xlsx')
        general_data: List[BranchInfo] = excel.general_data
        branch_data: List[BranchInfo] = excel.branch_data

        unique_branches: List[str] = self.unique_list([data.branch for data in branch_data])

        for period, functions in function_data[month].items():
            date_to: str = self.get_day(day=calendar.monthrange(today.year, month)[1], month=month, year=year)
            date_from: str = date_to if period == 'last_daily' else self.get_day(month=self.get_first_days(month)[period], year=year)

            for function in functions:
                if function in ['S_CLI_003', 'S_CLI_004', 'S_CLI_013', 'S_CLI_014']:
                    for branch in unique_branches:
                        self.info.append(BranchInfo(
                            branch=branch,
                            date_from=date_from,
                            date_to=date_to,
                            action=function,
                        ))
                    continue
                if function == 'Z_160_GL_003':
                    self.info.append(BranchInfo(
                        branch='00',
                        date_from=date_from,
                        date_to=date_to,
                        action=function,
                    ))
                    continue
                for data in general_data:
                    self.info.append(BranchInfo(
                        account=data.account,
                        branch=data.branch,
                        date_from=date_from,
                        date_to=date_to,
                        action=function,
                        account_name=data.account_name
                    ))
                for data in branch_data:
                    self.info.append(BranchInfo(
                        account=data.account,
                        branch=data.branch,
                        date_from=date_from,
                        date_to=date_to,
                        action=function,
                    ))

    @staticmethod
    def get_day(month: int, day: int = None, year: int = datetime.datetime.now().year) -> str:
        return datetime.datetime(
            year=year,
            month=month,
            day=day if day else 1
        ).strftime('%Y-%m-%d')

    @staticmethod
    def get_first_days(month) -> Dict:
        return {'monthly': month, 'quarterly': month - 2, 'six_monthly': month - 5, 'nine_monthly': month - 8, 'yearly': month - 11}

    @staticmethod
    def unique_list(_list) -> List[str]:
        seen: Set[Any] = set()
        seen_add = seen.add
        return [x for x in _list if not (x in seen or seen_add(x))]
